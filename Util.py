from os import listdir
from os.path import isfile, join
import pandas as pd
import openpyxl
from time import clock
import random
import math
import re

import Logic

class Utils:
    def __init__(self):
        self.ext_txt = ".txt"
        self.ext_dat = ".dat"
        self.ext_xlsx = ".xlsx"
        self.ext_fa = ".fa"

    """
    :return
        result_dict = {'chromosome': [[seq index from gene change, 'Cellline', 'Genome_Change', 'Gene', 'Genomesequence', 'cDNA_change', 'protein_change', 'Codon_Change']
                        ,'chr1': [[12726571, 'NCIH522', 'g.chr1:12726571G>A', 'AADACL4', 'Missense_Mutation', 'c.1049G>A', 'p.R350H', 'c.(1048-1050)cGt>cAt']
                            , [162824917, 'NCIH522', 'g.chr1:162824917C>G', 'C1orf110', 'Missense_Mutation', 'c.547G>C', 'p.E183Q', 'c.(547-549)Gag>Cag']
                            , [207195620, 'NCIH522', 'g.chr1:207195620T>C', 'C1orf116', 'Missense_Mutation', 'c.1489A>G', 'p.T497A', 'c.(1489-1491)Act>Gct']]
                        ,'chr2': [[158261913, 'NCIH522', 'g.chr1:158261913A>C', 'CD1C', 'Missense_Mutation', 'c.368A>C', 'p.H123P', 'c.(367-369)cAt>cCt']
                        , [158152749, 'NCIH522', 'g.chr1:158152749A>G', 'CD1D', 'Missense_Mutation', 'c.689A>G', 'p.Y230C', 'c.(688-690)tAc>tGc']
    """
    def read_txt_dvd_by_tab(self, path):
        result_dict = {}
        with open(path + self.ext_txt, "r") as f:
            head_list = f.readline().replace("\n", "").split("\t")
            while True:
                tmp_line = f.readline().replace("\n", "")
                if tmp_line == '':
                    break
                val_list = tmp_line.split("\t")

                gene_chng_list = val_list[1].replace("g.","").split(":")
                if gene_chng_list[0] in result_dict:
                    seq_key = [int(re.search(r'\d+', gene_chng_list[1].split("_")[0]).group())]
                    result_dict[gene_chng_list[0]].append(seq_key + val_list)
                else:
                    seq_key = [int(re.search(r'\d+', gene_chng_list[1].split("_")[0]).group())]
                    result_dict.update({gene_chng_list[0]: [seq_key + val_list]})

        return result_dict

    """
    :param
        chr_path : chromosome files path
        f_name : chr1
        max_len
    """
    def read_seq(self, chr_path, f_name, max_len):
        logic = Logic.Logics()
        with open(chr_path + f_name + self.ext_fa, "r") as f:
            header = f.readline()  # header ignored : >chr19
            print("header : " + header)

            idx = 1
            tmp_p_str = ""
            tmp_m_str = ""

            while True:
                c = f.read(1)
                if c == "":
                    break
                if "\n" in c:
                    continue
                elif "\r" in c:
                    continue

                # print(c + " " + str(idx))

                tmp_p_str = tmp_p_str + c.upper()
                tmp_m_str = tmp_m_str + logic.get_complementary(c.upper())

                if len(tmp_p_str) > max_len:
                    tmp_p_str = tmp_p_str[-max_len:]
                    tmp_m_str = tmp_m_str[-max_len:]

                if len(tmp_p_str) == max_len:
                    pass

                idx = idx + 1

    def make_excel(self, path, input_dict, out_dict):
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        row = 1
        sheet.cell(row=row, column=1, value="Cellline")
        sheet.cell(row=row, column=2, value='Genome_Change')
        sheet.cell(row=row, column=3, value='Genomesequence')
        sheet.cell(row=row, column=4, value='forward')
        sheet.cell(row=row, column=5, value='backward')
        sheet.cell(row=row, column=6, value='Strand')

        for chr_key, val_dict in input_dict.items():
            for val_arr in val_dict.values():
                for val_input in val_arr:
                    Cellline = val_input[0]
                    Genome_Change = val_input[1]
                    tmp_key = Cellline + "^" + Genome_Change
                    Genomesequence = val_input[3]
                    if tmp_key in out_dict[chr_key]:
                        val_out = out_dict[chr_key][tmp_key]
                        row = row + 1
                        sheet.cell(row=row, column=1, value=Cellline)
                        sheet.cell(row=row, column=2, value=Genome_Change)
                        sheet.cell(row=row, column=3, value=Genomesequence)
                        sheet.cell(row=row, column=4, value=val_out[2])
                        if len(val_out) > 3:
                            sheet.cell(row=row, column=5, value=val_out[3])
                        sheet.cell(row=row, column=6, value=val_out[0])
                        sheet.cell(row=row, column=7, value=len(val_out[2]))
                        if len(val_out) > 3:
                            sheet.cell(row=row, column=8, value=len(val_out[3]))
                    # else:
                        # print(val_input)

        workbook.save(filename=path + "_" + str(clock()) + self.ext_xlsx)
